import os
import re
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional, Any

from openpyxl import load_workbook


@dataclass
class ColumnInfo:
    name: str
    dtype: str
    attr: str


@dataclass
class TableInfo:
    db: str
    division: str
    table: str
    entity: str
    description: str
    columns: List[ColumnInfo]


def _norm(s: Optional[str]) -> str:
    return (str(s).strip() if s is not None else "")


def _canon(s: Optional[str]) -> str:
    return re.sub(r"[^a-z0-9_]+", "", (s or "").lower())


class SchemaRegistry:
    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path
        self.tables: List[TableInfo] = []
        self._index: List[Tuple[str, TableInfo]] = []

    def load(self) -> None:
        if not os.path.exists(self.xlsx_path):
            raise FileNotFoundError(f"Dictionary xlsx not found: {self.xlsx_path}")

        wb = load_workbook(self.xlsx_path, data_only=True)

        sh_table = wb["Table"]
        sh_col = wb["Column"]

        header = [c.value for c in next(sh_table.iter_rows(min_row=1, max_row=1))]
        col_map = {name: i for i, name in enumerate(header) if name}

        def v(row, key):
            idx = col_map.get(key)
            return row[idx].value if idx is not None else None

        table_rows: Dict[str, Dict[str, Any]] = {}
        for row in sh_table.iter_rows(min_row=2):
            db = v(row, "DB")
            div = v(row, "Division of work")
            tname = v(row, "Table Name")
            entity = v(row, "Entity Name") or ""
            desc = v(row, "Description") or ""
            if not db or not tname:
                continue

            dbs = _norm(db)
            tns = _norm(tname)
            key = f"{dbs}::{tns}"
            table_rows[key] = {
                "db": dbs,
                "division": _norm(div),
                "table": tns,
                "entity": _norm(entity),
                "description": _norm(desc),
                "columns": [],
            }

        header2 = [c.value for c in next(sh_col.iter_rows(min_row=1, max_row=1))]
        col_map2 = {name: i for i, name in enumerate(header2) if name}

        def v2(row, key):
            idx = col_map2.get(key)
            return row[idx].value if idx is not None else None

        for row in sh_col.iter_rows(min_row=2):
            db = v2(row, "DB")
            tname = v2(row, "Table Name")
            cname = v2(row, "Column Name")
            attr = v2(row, "Attribute Name") or ""
            dtype = v2(row, "Datatype") or ""

            if not tname or not cname:
                continue

            if db:
                key = f"{_norm(db)}::{_norm(tname)}"
            else:
                candidates = [k for k in table_rows.keys() if k.endswith(f"::{_norm(tname)}")]
                key = candidates[0] if candidates else None

            if not key or key not in table_rows:
                continue

            table_rows[key]["columns"].append(
                ColumnInfo(
                    name=_norm(cname),
                    dtype=_norm(dtype),
                    attr=_norm(attr),
                )
            )

        self.tables = [TableInfo(**t) for t in table_rows.values()]

        self._index = []
        for t in self.tables:
            blob = " ".join(
                [
                    t.db,
                    t.division,
                    t.table,
                    t.entity,
                    t.description,
                    " ".join([c.name for c in t.columns]),
                    " ".join([c.attr for c in t.columns]),
                ]
            ).lower()
            self._index.append((blob, t))

    def search(self, query: str, top_k: int = 8) -> List[TableInfo]:
        q = (query or "").lower().strip()
        if not q:
            return []

        tokens = [x for x in re.split(r"[^a-z0-9_]+", q) if x]
        scored: List[Tuple[int, TableInfo]] = []

        for blob, t in self._index:
            score = 0

            for tok in tokens:
                if tok in blob:
                    score += 2

            if t.table.lower() in q:
                score += 10

            role = self.infer_table_role(t)
            if "sales" in q and role == "sales_fact":
                score += 20
            if any(x in q for x in ["store", "салбар", "дэлгүүр"]) and role == "store_dimension":
                score += 14
            if any(x in q for x in ["product", "бараа", "бүтээгдэхүүн", "item"]) and role == "product_dimension":
                score += 14
            if any(x in q for x in ["promotion", "promo", "event", "campaign", "хямдрал"]) and role in {
                "event_dimension",
                "event_goods_dimension",
            }:
                score += 14
            if any(x in q for x in ["stock", "inventory", "үлдэгдэл", "агуулах"]) and role == "inventory_fact":
                score += 14

            scored.append((score, t))

        scored.sort(key=lambda x: x[0], reverse=True)
        return [t for s, t in scored if s > 0][:top_k]

    def highlights(self, t: TableInfo) -> Dict[str, List[str]]:
        cols = [c.name for c in t.columns]
        lc = [x.lower() for x in cols]

        date_cols = [
            cols[i]
            for i, x in enumerate(lc)
            if "date" in x or x.endswith("_dt") or x.endswith("_ymd") or x.endswith("_dtm")
        ]
        store_cols = [
            cols[i]
            for i, x in enumerate(lc)
            if x in ("storeid", "store_id", "bizloc_cd", "bizloc_org_cd", "org_cd", "store_no")
        ]
        metric_cols = [
            cols[i]
            for i, x in enumerate(lc)
            if x
               in (
                   "netsale",
                   "grosssale",
                   "tax_vat",
                   "discount",
                   "actualcost",
                   "soldqty",
                   "qty",
                   "vat",
                   "city_tax",
                   "stockqty",
                   "stck_qty",
                   "amount",
                   "stockamt",
                   "cost",
                   "total_amount",
               )
        ]
        key_cols = [
            cols[i]
            for i, x in enumerate(lc)
            if x
               in (
                   "gds_cd",
                   "item_cd",
                   "promotionid",
                   "evt_cd",
                   "receiptno",
                   "cate_cd",
                   "storeid",
                   "bizloc_cd",
                   "store",
                   "item",
                   "seq_no",
                   "plu_cd",
               )
        ]
        name_cols = [
            cols[i]
            for i, x in enumerate(lc)
            if x
               in (
                   "gds_nm",
                   "item_nm",
                   "store_nm",
                   "storename",
                   "bizloc_nm",
                   "cate_nm",
                   "brand_nm",
                   "gds_label_nm",
                   "evt_nm",
                   "evt_short_nm",
               )
        ]

        return {
            "date_cols": date_cols[:10],
            "store_cols": store_cols[:10],
            "metric_cols": metric_cols[:20],
            "key_cols": key_cols[:20],
            "name_cols": name_cols[:20],
        }

    def infer_table_role(self, t: TableInfo) -> str:
        name = (t.table or "").lower()
        entity = (t.entity or "").lower()
        desc = (t.description or "").lower()
        cols = {c.name.lower() for c in t.columns}

        if name == "cluster_main_sales":
            return "sales_fact"

        if name in {"dimension_im"}:
            return "product_dimension"

        if name in {"dimension_sm"}:
            return "store_dimension"

        if name in {"dimension_lem"}:
            return "event_dimension"

        if name in {"dimension_leg"}:
            return "event_goods_dimension"

        if name.startswith("agg_sales_") or name in {"agg_item_sales", "main_sales", "main_sales_mv"}:
            return "sales_aggregate"

        if "sales" in name and {"netsale", "grosssale"} & cols:
            return "sales_fact"

        if "stock" in name or {"stockqty", "stck_qty", "stockamt"} & cols:
            return "inventory_fact"

        if "item master" in entity or "item master" in desc:
            return "product_dimension"

        if "store master" in entity or "store master" in desc:
            return "store_dimension"

        if "event master" in entity or "event master" in desc:
            return "event_dimension"

        if "event goods master" in entity or "event goods master" in desc:
            return "event_goods_dimension"

        return "unknown"

    def to_table_card(self, t: TableInfo, max_cols: int = 80) -> Dict[str, Any]:
        h = self.highlights(t)
        return {
            "db": t.db,
            "table": t.table,
            "role": self.infer_table_role(t),
            "entity": t.entity,
            "description": t.description,
            "highlights": h,
            "columns": [{"name": c.name, "type": c.dtype, "attr": c.attr} for c in t.columns[:max_cols]],
        }

    def build_relationships(self) -> List[Dict[str, Any]]:
        rel: List[Dict[str, Any]] = []

        # High-confidence manual joins
        rel.append(
            {
                "left": "Cluster_Main_Sales.GDS_CD",
                "right": "Dimension_IM.GDS_CD",
                "type": "join_key",
                "label": "product",
                "score": 1000,
            }
        )
        rel.append(
            {
                "left": "Cluster_Main_Sales.StoreID",
                "right": "Dimension_SM.BIZLOC_CD",
                "type": "join_key",
                "label": "store",
                "score": 1000,
            }
        )
        rel.append(
            {
                "left": "Cluster_Main_Sales.PromotionID",
                "right": "Dimension_LEM.EVT_CD",
                "type": "join_key",
                "label": "promotion_event",
                "score": 980,
            }
        )
        rel.append(
            {
                "left": "Dimension_LEG.EVT_CD",
                "right": "Dimension_LEM.EVT_CD",
                "type": "join_key",
                "label": "event_goods_to_event",
                "score": 960,
            }
        )

        rel.append(
            {
                "table": "Dimension_IM",
                "name_column": "GDS_NM",
                "type": "name_column",
                "label": "product_name",
                "score": 1000,
            }
        )
        rel.append(
            {
                "table": "Dimension_SM",
                "name_column": "BIZLOC_NM",
                "type": "name_column",
                "label": "store_name",
                "score": 1000,
            }
        )
        rel.append(
            {
                "table": "Dimension_LEM",
                "name_column": "EVT_NM",
                "type": "name_column",
                "label": "event_name",
                "score": 950,
            }
        )

        tbl_cols: Dict[str, List[Dict[str, str]]] = {}
        for t in self.tables:
            tbl_cols[t.table] = [
                {
                    "name": c.name,
                    "attr": (c.attr or "").lower(),
                    "canon": _canon(c.name),
                }
                for c in t.columns
            ]

        join_canon = {
            "product": {"gds_cd", "item_cd"},
            "store": {"storeid", "store_id", "bizloc_cd", "org_cd", "store_no"},
            "category": {"cate_cd"},
            "receipt": {"receiptno"},
            "promotion": {"promotionid", "evt_cd"},
        }

        for group, canon_set in join_canon.items():
            occ: List[Tuple[str, str]] = []
            for tbl, cols in tbl_cols.items():
                for c in cols:
                    if c["canon"] in canon_set:
                        occ.append((tbl, c["name"]))

            for i in range(len(occ)):
                for j in range(i + 1, len(occ)):
                    lt, lc = occ[i]
                    rt, rc = occ[j]
                    if lt == rt:
                        continue
                    rel.append(
                        {
                            "left": f"{lt}.{lc}",
                            "right": f"{rt}.{rc}",
                            "type": "join_key",
                            "label": group,
                            "score": 120 if lc == rc else 80,
                        }
                    )

        name_candidates = {
            "Dimension_IM": ["GDS_NM", "GDS_LABEL_NM"],
            "Dimension_SM": ["BIZLOC_NM", "StoreName", "STORE_NM"],
            "Dimension_LEM": ["EVT_NM", "EVT_SHORT_NM"],
        }

        table_map = {t.table: t for t in self.tables}
        for tbl, candidates in name_candidates.items():
            t = table_map.get(tbl)
            if not t:
                continue
            existing = {c.name for c in t.columns}
            for name_col in candidates:
                if name_col in existing:
                    rel.append(
                        {
                            "table": tbl,
                            "name_column": name_col,
                            "type": "name_column",
                            "label": f"{tbl} name",
                            "score": 200,
                        }
                    )
                    break

        rel.sort(key=lambda x: x.get("score", 0), reverse=True)

        dedup = []
        seen = set()
        for r in rel:
            key = tuple(sorted([(k, str(v)) for k, v in r.items()]))
            if key in seen:
                continue
            seen.add(key)
            dedup.append(r)

        return dedup
